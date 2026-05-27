"""
Predator Protocol — Markets Excel Ingest (v2)
=============================================
Reads Mega_Markets_Historical.xlsx (and optionally Markets_1_.xlsx when it
exists) and emits docs/data/market_returns.json in the §2.3 canonical
contract shape.

§2.3 contract shape:
{
  "asof": "YYYY-MM-DD",
  "generated_utc": "...",
  "assets": {
    "GOLD": {
      "name": "Gold",
      "category": "precious_metals",   // lowercase_snake enum
      "native_ccy": "USD",
      "source": "Mega_Markets_Historical.xlsx:Precious_Metals",
      "first": "1833-01",
      "last":  "2026-05",
      "monthly": [["1833-01", 18.93], ...]   // [YYYY-MM, EOP price]
    }
  },
  "fx": {
    "USDINR": [["1973-01", 8.02], ...],
    "USDJPY": [...],
    "EURUSD": [...],
    "GBPUSD": [...],
    "USDCNY": [...]
  },
  "cpi": {
    "US": [["1947-01", 21.48], ...]
  },
  "rates": {
    "us_10y": { "name": "US 10Y Treasury", "values": [["1953-04", 2.83], ...] },
    ...
  },
  "events": [
    {"year": 2008, "label": "GFC"},
    ...
  ]
}

Key design decisions:
- Years are strings everywhere (YYYY-MM) — kills the year-key type mismatch bug.
- Returns are computed from prices client-side, never stored.
- monthly = sorted list of [YYYY-MM, price] pairs (not a dict) for deterministic
  ordering and smaller JSON (no repeated key overhead).
- fx, cpi, rates are top-level sections so the currency/real engine can find them
  without scanning the assets dict.
- The existing ingest_mega_xl.py is preserved for backward compat; this module
  supersedes it for the new dashboard.

Usage:
    python -m predator.ingest_markets_xl                    # full rebuild
    python -m predator.ingest_markets_xl --dry-run          # print plan, no write
    python -m predator.ingest_markets_xl --merge-existing   # merge with existing JSON
    python -m predator.ingest_markets_xl --check-freshness  # freshness gate only
"""
from __future__ import annotations

import argparse
import json
import math
import os
import sys
from datetime import datetime, timezone, date
from pathlib import Path
from typing import Any

import pandas as pd

# ─── Platform encoding fix (Windows CI) ──────────────────────────────────────
if sys.platform.startswith("win"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except AttributeError:
        pass


# ─── Paths ───────────────────────────────────────────────────────────────────

REPO_ROOT    = Path(__file__).resolve().parent.parent
MEGA_XL      = REPO_ROOT / "Mega_Markets_Historical.xlsx"
MARKETS1_XL  = REPO_ROOT / "Markets_1_.xlsx"   # optional; used when present
OUTPUT_PATH  = REPO_ROOT / "docs" / "data" / "market_returns.json"


# ─── Asset registry ──────────────────────────────────────────────────────────
# Maps (sheet_name, Index_Name) → (asset_id, category, native_ccy, display_name)
# This is the single source of truth for asset identity — never drive grouping
# off display strings.

ASSET_REGISTRY: dict[tuple[str, str], tuple[str, str, str, str]] = {
    # (sheet, Index_Name): (asset_id, category, native_ccy, display_name)
    # ── Equities ──────────────────────────────────────────────────────────
    ("Equities", "S&P 500"):                              ("sp500",       "equity",          "USD", "S&P 500"),
    ("Equities", "Dow Jones Industrial Average"):         ("djia",        "equity",          "USD", "Dow Jones"),
    ("Equities", "NASDAQ Composite"):                     ("nasdaq",      "equity",          "USD", "NASDAQ Composite"),
    ("Equities", "NASDAQ-100"):                           ("nasdaq100",   "equity",          "USD", "NASDAQ-100"),
    ("Equities", "Nikkei 225"):                           ("nikkei",      "equity",          "JPY", "Nikkei 225"),
    ("Equities", "BSE Sensex"):                           ("sensex",      "equity",          "INR", "BSE Sensex"),
    ("Equities", "DAX (Curvo synthetic TR, base 10000)"): ("dax",         "equity",          "EUR", "DAX (Germany)"),
    # ── International Equities ────────────────────────────────────────────
    ("International_Equities", "FTSE 100 (GBP)"):         ("ftse100",     "equity",          "GBP", "FTSE 100"),
    ("International_Equities", "Hang Seng (HKD)"):        ("hang_seng",   "equity",          "HKD", "Hang Seng"),
    ("International_Equities", "ASX 200 (AUD)"):          ("asx200",      "equity",          "AUD", "ASX 200"),
    ("International_Equities", "Bovespa (BRL)"):          ("bovespa",     "equity",          "BRL", "Bovespa"),
    ("International_Equities", "Shanghai Composite (CNY)"): ("shanghai",  "equity",          "CNY", "Shanghai Composite"),
    # ── Precious Metals ───────────────────────────────────────────────────
    ("Precious_Metals", "Gold (USD/oz)"):                 ("gold",        "precious_metals", "USD", "Gold"),
    ("Precious_Metals", "Silver (USD/oz)"):               ("silver",      "precious_metals", "USD", "Silver"),
    ("Precious_Metals", "Platinum (USD/oz)"):             ("platinum",    "precious_metals", "USD", "Platinum"),
    ("Precious_Metals", "Palladium (USD/oz)"):            ("palladium",   "precious_metals", "USD", "Palladium"),
    # ── Commodities ───────────────────────────────────────────────────────
    ("Commodities", "Copper (USD/lb)"):                   ("copper_lb",   "base_metals",     "USD", "Copper"),
    ("Commodities", "Corn (USD/bushel)"):                 ("corn",        "agriculture",     "USD", "Corn"),
    ("Commodities", "Wheat (USD/bushel)"):                ("wheat",       "agriculture",     "USD", "Wheat"),
    ("Commodities", "Soybeans (USD/bushel)"):             ("soybeans",    "agriculture",     "USD", "Soybeans"),
    ("Commodities", "Sugar (USD/lb)"):                    ("sugar",       "agriculture",     "USD", "Sugar"),
    ("Commodities", "Coffee Arabica (USD/lb)"):           ("coffee",      "agriculture",     "USD", "Coffee Arabica"),
    ("Commodities", "Cotton (USD/lb)"):                   ("cotton",      "agriculture",     "USD", "Cotton"),
    # ── Energy ────────────────────────────────────────────────────────────
    ("Energy", "WTI Crude Oil (USD/bbl)"):                ("wti_crude",   "energy",          "USD", "WTI Crude Oil"),
    ("Energy", "Brent Crude Oil (USD/bbl)"):              ("brent_crude", "energy",          "USD", "Brent Crude Oil"),
}

# FX series: sheet → (fx_pair_id, description)
# Stored as USD per 1 unit of foreign currency (i.e. USDINR = USD/INR rate)
FX_REGISTRY: dict[tuple[str, str], tuple[str, str]] = {
    # The Mega XL doesn't have a dedicated FX sheet; FX comes from markets_history.py
    # This registry is here for Markets_1_.xlsx when it has a USD_INR sheet.
}

# Crisis events for annotation
EVENTS: list[dict] = [
    {"year": 1973, "label": "Oil embargo"},
    {"year": 1979, "label": "Volcker shock"},
    {"year": 1987, "label": "Black Monday"},
    {"year": 1997, "label": "Asian crisis"},
    {"year": 1998, "label": "LTCM / Russia"},
    {"year": 2000, "label": "Dot-com peak"},
    {"year": 2001, "label": "9/11"},
    {"year": 2008, "label": "GFC"},
    {"year": 2011, "label": "Euro debt crisis"},
    {"year": 2015, "label": "China devaluation"},
    {"year": 2020, "label": "COVID-19"},
    {"year": 2022, "label": "Rate shock"},
]

# Freshness cadence: how many calendar days old is acceptable for each category
# Monthly series: up to 35 days (one month + buffer)
# Daily series: up to 4 days (weekend + 1 buffer)
FRESHNESS_CADENCE: dict[str, int] = {
    "equity":          35,
    "precious_metals": 35,
    "energy":          35,
    "base_metals":     35,
    "agriculture":     35,
    "fx":              35,
    "real_estate":     35,
    "rates":           35,
}


# ─── Safe JSON encoder ────────────────────────────────────────────────────────

class _SafeEncoder(json.JSONEncoder):
    def iterencode(self, o, _one_shot=False):
        return super().iterencode(self._sanitise(o), _one_shot)

    def _sanitise(self, obj):
        if isinstance(obj, float):
            return None if (math.isnan(obj) or math.isinf(obj)) else obj
        if isinstance(obj, dict):
            return {k: self._sanitise(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [self._sanitise(v) for v in obj]
        return obj


def _write_json_atomic(path: Path, text: str) -> None:
    """Write JSON atomically: write to .tmp then os.replace to avoid truncated files."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(text, encoding="utf-8")
    os.replace(tmp, path)


def _dumps(obj: Any, **kwargs) -> str:
    kwargs.setdefault("cls", _SafeEncoder)
    return json.dumps(obj, **kwargs)


# ─── Git LFS pointer detection ───────────────────────────────────────────────
# When the repo is cloned without `git lfs install`/`git lfs pull`, when the LFS
# server is unreachable, or when the LFS bandwidth quota is exhausted, the
# `Mega_Markets_Historical.xlsx` working-tree file is a small text "pointer"
# rather than the real binary blob. Letting `pandas.read_excel` see that file
# produces a confusing `BadZipFile` error; detecting the pointer up front lets
# us raise a clear, actionable LFS-specific message instead.

_LFS_POINTER_HEADER = b"version https://git-lfs.github.com/spec/"
_LFS_POINTER_MAX_BYTES = 1024


def _is_lfs_pointer(path: Path) -> bool:
    """Return True if ``path`` is a Git LFS pointer file rather than the real binary.

    A Git LFS pointer is a tiny text stub (≤1024 bytes) whose first line begins
    with ``version https://git-lfs.github.com/spec/``. This helper performs a
    fast size + magic-bytes check without parsing the rest of the file.
    """
    try:
        if path.stat().st_size > _LFS_POINTER_MAX_BYTES:
            return False
    except OSError:
        return False
    try:
        head = path.read_bytes()[:64]
    except OSError:
        return False
    return head.startswith(_LFS_POINTER_HEADER)


# ─── Excel reader ─────────────────────────────────────────────────────────────

def _read_mega_xl(xl_path: Path, asset_filter: set[str] | None = None) -> dict[str, pd.DataFrame]:
    """
    Read Mega_Markets_Historical.xlsx.

    Returns {asset_id: DataFrame(Date, Close)} for all registered assets.
    Locates blocks by header text (not fixed column index) for robustness.

    Args:
        xl_path: Path to the Excel file.
        asset_filter: If provided, only return these asset IDs (e.g. {'sp500', 'gold'}).
    """
    print(f"\nReading: {xl_path.name}")

    # Fail fast with a clear LFS-specific error before invoking pandas.read_excel.
    # This is the single error path for clone-without-LFS, LFS-server-unreachable,
    # and LFS-quota-exhausted scenarios — all leave a pointer file in place.
    if _is_lfs_pointer(xl_path):
        raise RuntimeError(
            "Mega_Markets_Historical.xlsx is a Git LFS pointer, not a real Excel file. "
            "Run 'git lfs install && git lfs pull'."
        )

    results: dict[str, pd.DataFrame] = {}

    # Build lookup: (sheet, index_name) → asset_id
    lookup = {(sheet, idx): (aid, cat, ccy, name)
              for (sheet, idx), (aid, cat, ccy, name) in ASSET_REGISTRY.items()}

    # Apply asset filter if provided
    if asset_filter:
        lookup = {k: v for k, v in lookup.items() if v[0] in asset_filter}

    sheets_to_read = set(sheet for (sheet, _) in lookup.keys())

    for sheet in sheets_to_read:
        try:
            df = pd.read_excel(xl_path, sheet_name=sheet)
        except Exception as e:
            print(f"  WARNING: Could not read sheet '{sheet}': {e}")
            continue

        # Locate the Date and Close columns robustly
        cols_lower = {c: c for c in df.columns}
        date_col = next((c for c in df.columns if str(c).strip().lower() == "date"), None)
        close_col = next((c for c in df.columns if str(c).strip().lower() == "close"), None)
        name_col = next((c for c in df.columns if str(c).strip().lower() in ("index_name", "name")), None)

        if date_col is None or close_col is None:
            print(f"  WARNING: Sheet '{sheet}' missing Date/Close columns. Got: {list(df.columns)}")
            continue

        if name_col is None:
            print(f"  WARNING: Sheet '{sheet}' missing Index_Name/Name column — cannot identify series. Got: {list(df.columns)}")
            continue

        for index_name in df[name_col].unique() if name_col else []:
            key = (sheet, str(index_name).strip())
            if key not in lookup:
                continue
            aid, cat, ccy, display = lookup[key]

            sub = df[df[name_col] == index_name][[date_col, close_col]].copy()
            sub.columns = ["Date", "Close"]
            sub["Date"] = pd.to_datetime(sub["Date"], errors="coerce")
            sub = sub.dropna(subset=["Date", "Close"])
            sub = sub[sub["Close"] > 0]
            sub = sub.sort_values("Date")

            results[aid] = sub
            print(f"  {sheet:25s} | {index_name:50s} → {aid} ({len(sub)} rows, "
                  f"{sub['Date'].min().strftime('%Y-%m')} → {sub['Date'].max().strftime('%Y-%m')})")

    return results


def _to_monthly_eop(df: pd.DataFrame) -> list[list]:
    """
    Resample daily data to monthly end-of-period close.

    Returns sorted list of [YYYY-MM, value] pairs.
    Years are strings — kills the year-key type mismatch bug permanently.
    """
    if df.empty:
        return []
    monthly = df.resample("ME", on="Date")["Close"].last().dropna()
    result = []
    for dt, val in monthly.items():
        ym = pd.Timestamp(dt).strftime("%Y-%m")
        result.append([ym, round(float(val), 4)])
    return sorted(result, key=lambda x: x[0])


# ─── Merge with existing JSON ─────────────────────────────────────────────────

def _load_existing(path: Path) -> dict[str, Any]:
    """Load existing market_returns.json. Returns empty structure on failure."""
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"  WARNING: Could not parse existing JSON ({e}). Starting fresh.")
    return {"asof": "", "generated_utc": "", "assets": {}, "fx": {}, "cpi": {}, "rates": {}, "events": EVENTS}


# Public aliases (used by markets_history.py and tests)
load_existing = _load_existing


def _merge_monthly(existing: list[list], new: list[list]) -> list[list]:
    """
    Merge two sorted [YYYY-MM, value] lists.
    New data takes priority for overlapping months.
    Returns sorted merged list.
    """
    merged: dict[str, float] = {}
    for ym, val in existing:
        merged[ym] = val
    for ym, val in new:
        merged[ym] = val  # new overwrites existing
    return sorted([[ym, v] for ym, v in merged.items()], key=lambda x: x[0])


# Public alias
merge_monthly = _merge_monthly


# ─── Freshness gate ───────────────────────────────────────────────────────────

def check_freshness(output: dict[str, Any], fail_on_stale: bool = True) -> list[str]:
    """
    Check that each asset's last data point is within its expected cadence.

    Returns list of stale series descriptions.
    If fail_on_stale=True, exits with code 1 when any series is stale.
    """
    today = date.today()
    stale: list[str] = []

    assets = output.get("assets", {})
    for asset_id, asset_data in assets.items():
        last_str = asset_data.get("meta", {}).get("last", asset_data.get("last", ""))
        if not last_str:
            stale.append(f"  STALE: {asset_id} — no 'last' date")
            continue
        try:
            # last is YYYY-MM; use last day of that month as the data date
            last_dt = pd.Timestamp(last_str + "-01") + pd.offsets.MonthEnd(0)
            last_date = last_dt.date()
        except Exception:
            stale.append(f"  STALE: {asset_id} — unparseable last date '{last_str}'")
            continue

        category = asset_data.get("meta", {}).get("category", asset_data.get("category", "equity"))
        max_age = FRESHNESS_CADENCE.get(category, 35)
        age_days = (today - last_date).days

        if age_days > max_age:
            name = asset_data.get("meta", {}).get("name", asset_data.get("name", asset_id))
            stale.append(
                f"  STALE: {asset_id} ({name}) — "
                f"last={last_str}, age={age_days}d > {max_age}d cadence"
            )

    if stale:
        print("\n⚠️  FRESHNESS GATE — stale series detected:")
        for s in stale:
            print(s)
        if fail_on_stale:
            print("\n❌ Build failed: stale data. Re-run the data pipeline.")
            sys.exit(1)
        else:
            print("\n⚠️  Continuing despite stale data (--no-fail-on-stale).")
    else:
        print("\n✅ Freshness gate passed — all series within cadence.")

    return stale


# ─── Build output ─────────────────────────────────────────────────────────────

def build_output(
    raw_series: dict[str, pd.DataFrame],
    existing: dict[str, Any],
    merge: bool = False,
) -> dict[str, Any]:
    """
    Build the §2.3 contract JSON from raw series DataFrames.

    If merge=True, merges with existing JSON (new data takes priority).
    """
    # Build asset lookup: asset_id → (cat, ccy, display, sheet)
    id_to_meta: dict[str, tuple[str, str, str, str]] = {}
    for (sheet, idx), (aid, cat, ccy, display) in ASSET_REGISTRY.items():
        id_to_meta[aid] = (cat, ccy, display, sheet)

    assets_out: dict[str, Any] = {}

    for asset_id, df in raw_series.items():
        meta = id_to_meta.get(asset_id)
        if meta is None:
            continue
        cat, ccy, display, sheet = meta

        new_monthly = _to_monthly_eop(df)
        if not new_monthly:
            print(f"  SKIP {asset_id}: no monthly data after resampling")
            continue

        if merge and asset_id in existing.get("assets", {}):
            existing_monthly = existing["assets"][asset_id].get("monthly", [])
            # Convert old "close" dict format if present
            if not existing_monthly and "close" in existing.get("assets", {}).get(asset_id, {}):
                existing_monthly = sorted(
                    [[ym, v] for ym, v in existing["assets"][asset_id]["close"].items()],
                    key=lambda x: x[0]
                )
            monthly = _merge_monthly(existing_monthly, new_monthly)
            print(f"  MERGE  {asset_id:15s}: {len(existing_monthly)} → {len(monthly)} months")
        else:
            monthly = new_monthly
            print(f"  NEW    {asset_id:15s}: {len(monthly)} months "
                  f"({monthly[0][0]} → {monthly[-1][0]})")

        # ── Source-label honesty (charter v2 Part 2) ─────────────────────
        # Excel is BACKFILL ONLY. If a previous step (markets_history.py)
        # already attached a live FRED/yfinance source, never overwrite it
        # with the Excel label — that's how the live JSON ended up claiming
        # "Mega_Markets_Historical.xlsx" for SP500 etc. when the data was
        # actually being extended monthly by yfinance.
        existing_source = (existing.get("assets", {}).get(asset_id, {})
                                                  .get("meta", {}).get("source", ""))
        excel_source = f"{MEGA_XL.name}:{sheet}"
        if existing_source and not existing_source.startswith(MEGA_XL.name):
            # Live source already set (e.g. "yfinance:^GSPC", "fred:DCOILWTICO").
            # Keep it; the Excel run is just supplying deep-history backfill.
            source_label = existing_source
        else:
            source_label = excel_source

        assets_out[asset_id] = {
            "meta": {
                "name":       display,
                "category":   cat,
                "native_ccy": ccy,
                "source":     source_label,
                "first":      monthly[0][0],
                "last":       monthly[-1][0],
            },
            "monthly":    monthly,
        }

    # Preserve existing fx, cpi, rates sections (populated by markets_history.py)
    fx_out    = existing.get("fx", {})
    cpi_out   = existing.get("cpi", {})
    rates_out = existing.get("rates", {})

    # Compute asof = latest last date across all assets AND fx/cpi/rates sections
    all_lasts = [v.get("meta", {}).get("last", v.get("last", "")) for v in assets_out.values() if v.get("meta", {}).get("last", v.get("last", ""))]
    # Include fx section lasts
    for pair_data in fx_out.values():
        if pair_data:
            all_lasts.append(pair_data[-1][0])
    # Include cpi section lasts
    for cpi_data in cpi_out.values():
        if cpi_data:
            all_lasts.append(cpi_data[-1][0])
    # Include rates section lasts
    for rate_data in rates_out.values():
        if isinstance(rate_data, dict) and rate_data.get("meta", {}).get("last"):
            all_lasts.append(rate_data["meta"]["last"])
    asof = max(all_lasts) if all_lasts else date.today().strftime("%Y-%m")

    return {
        "asof":          asof,
        "generated_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "assets":        assets_out,
        "fx":            fx_out,
        "cpi":           cpi_out,
        "rates":         rates_out,
        "events":        EVENTS,
    }


# ─── Main ─────────────────────────────────────────────────────────────────────

def process(
    dry_run: bool = False,
    merge: bool = False,
    check_freshness_only: bool = False,
    fail_on_stale: bool = True,
    asset_filter: set[str] | None = None,
) -> dict[str, Any]:
    """
    Full pipeline: read Excel → build §2.3 contract → write JSON.

    Args:
        dry_run: Print plan without writing.
        merge: Merge with existing JSON (new data takes priority).
        check_freshness_only: Only run the freshness gate.
        fail_on_stale: Exit 1 if stale data detected.
        asset_filter: If provided, only process these asset IDs (e.g. {'sp500', 'gold'}).
    """
    if check_freshness_only:
        if not OUTPUT_PATH.exists():
            print(f"ERROR: {OUTPUT_PATH} not found. Run without --check-freshness first.")
            sys.exit(1)
        existing = _load_existing(OUTPUT_PATH)
        check_freshness(existing, fail_on_stale=fail_on_stale)
        return existing

    # Determine which Excel to read
    xl_path = MEGA_XL
    if MARKETS1_XL.exists():
        # Validate Markets_1_.xlsx has at least one expected sheet before switching
        try:
            import openpyxl
            wb = openpyxl.load_workbook(MARKETS1_XL, read_only=True, data_only=True)
            expected_sheets = set(sheet for (sheet, _) in ASSET_REGISTRY.keys())
            found_sheets = set(wb.sheetnames) & expected_sheets
            wb.close()
            if found_sheets:
                print(f"  Found Markets_1_.xlsx with sheets {found_sheets} — will use it as primary source")
                xl_path = MARKETS1_XL
            else:
                print(f"  WARNING: Markets_1_.xlsx has no expected sheets (got {wb.sheetnames}) — falling back to {MEGA_XL.name}")
        except Exception as e:
            print(f"  WARNING: Could not inspect Markets_1_.xlsx ({e}) — falling back to {MEGA_XL.name}")

    if not xl_path.exists():
        print(f"ERROR: Excel file not found: {xl_path}")
        sys.exit(1)

    # Read Excel
    raw_series = _read_mega_xl(xl_path, asset_filter=asset_filter)

    if not raw_series:
        print("WARNING: No matching series found in Excel.")
        return {}

    # Load existing JSON for merge
    # ALWAYS load existing to preserve fx/cpi/rates (populated by markets_history.py).
    # The `merge` flag only controls whether existing asset monthly data is merged.
    existing = _load_existing(OUTPUT_PATH)

    print(f"\nBuilding §2.3 contract from {len(raw_series)} series...")
    output = build_output(raw_series, existing, merge=merge)

    if dry_run:
        print("\n--dry-run: output NOT written.")
        for key in list(raw_series.keys())[:3]:
            if key in output["assets"]:
                m = output["assets"][key]["monthly"]
                print(f"\n  {key}: {len(m)} months, {m[0][0]} → {m[-1][0]}")
                print(f"    first 3: {m[:3]}")
                print(f"    last  3: {m[-3:]}")
        return output

    # Write output
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    _write_json_atomic(OUTPUT_PATH, _dumps(output, separators=(",", ":")))
    size_kb = OUTPUT_PATH.stat().st_size / 1024
    print(f"\nWrote: {OUTPUT_PATH} ({size_kb:.1f} KB)")
    print(f"  asof: {output['asof']}")
    print(f"  assets: {len(output['assets'])}")

    # Run freshness gate after writing
    check_freshness(output, fail_on_stale=fail_on_stale)

    return output


# ─── CLI ─────────────────────────────────────────────────────────────────────

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="predator.ingest_markets_xl",
        description="Ingest Markets Excel → docs/data/market_returns.json (§2.3 contract)",
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Print plan without writing.")
    parser.add_argument("--merge-existing", action="store_true",
                        help="Merge with existing JSON (new data takes priority).")
    parser.add_argument("--check-freshness", action="store_true",
                        help="Only run the freshness gate on the existing JSON.")
    parser.add_argument("--no-fail-on-stale", action="store_true",
                        help="Warn on stale data instead of failing the build.")
    parser.add_argument("--assets", type=str, default=None,
                        help="Comma-separated asset keys to process (e.g. sp500,gold). Default: all.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    print("═" * 60)
    print("  Predator Protocol — Markets Excel Ingest v2")
    print("═" * 60)

    # Parse --assets filter if provided
    asset_filter: set[str] | None = None
    if args.assets:
        asset_filter = {a.strip().lower() for a in args.assets.split(",") if a.strip()}

    process(
        dry_run=args.dry_run,
        merge=args.merge_existing,
        check_freshness_only=args.check_freshness,
        fail_on_stale=not args.no_fail_on_stale,
        asset_filter=asset_filter,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
